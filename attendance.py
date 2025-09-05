import serial
import openpyxl
from datetime import datetime

# Open serial port (adjust COM port for Windows or /dev/ttyUSBx for Linux/Mac)
ser = serial.Serial("COM3", 9600)

# Create workbook with separate sheets
try:
    workbook = openpyxl.load_workbook("Attendance.xlsx")
except:
    workbook = openpyxl.Workbook()
    workbook.create_sheet("RFID", 0)
    workbook.create_sheet("Ultrasonic", 1)
    workbook.create_sheet("Face", 2)
    workbook.save("Attendance.xlsx")

sheet_rfid = workbook["RFID"]
sheet_ultrasonic = workbook["Ultrasonic"]

# Add headers if empty
if sheet_rfid.max_row == 1:
    sheet_rfid.append(["RFID_UID", "Timestamp"])
if sheet_ultrasonic.max_row == 1:
    sheet_ultrasonic.append(["Status", "Timestamp"])

print("Listening for Arduino...")

while True:
    data = ser.readline().decode().strip()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if data.startswith("RFID:"):
        uid = data.split(":")[1]
        sheet_rfid.append([uid, timestamp])
        print(f"[RFID] {uid} at {timestamp}")

    elif data.startswith("ULTRASONIC:"):
        sheet_ultrasonic.append([data, timestamp])
        print(f"[ULTRASONIC] {data} at {timestamp}")

    workbook.save("Attendance.xlsx")
