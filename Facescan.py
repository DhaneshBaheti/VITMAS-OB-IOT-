import cv2
import face_recognition
import os
import openpyxl
from datetime import datetime

# Load Excel
try:
    workbook = openpyxl.load_workbook("Attendance.xlsx")
except:
    workbook = openpyxl.Workbook()
    workbook.create_sheet("Face", 0)
    workbook.save("Attendance.xlsx")

sheet_face = workbook["Face"]

if sheet_face.max_row == 1:
    sheet_face.append(["Name", "Timestamp"])

# Load known faces
known_encodings = []
known_names = []

for filename in os.listdir("faces"):
    if filename.endswith(".jpg") or filename.endswith(".png"):
        img = face_recognition.load_image_file(f"faces/{filename}")
        encoding = face_recognition.face_encodings(img)[0]
        known_encodings.append(encoding)
        known_names.append(filename.split(".")[0])  # use filename as name

# Start webcam
cap = cv2.VideoCapture(0)

print("Face Attendance Running...")

while True:
    ret, frame = cap.read()
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for encoding in face_encodings:
        matches = face_recognition.compare_faces(known_encodings, encoding)
        if True in matches:
            idx = matches.index(True)
            name = known_names[idx]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            sheet_face.append([name, timestamp])
            workbook.save("Attendance.xlsx")

            print(f"[FACE] {name} at {timestamp}")

    cv2.imshow("Face Attendance", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
